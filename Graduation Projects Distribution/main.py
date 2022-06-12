import random
import openpyxl

#Generate 36 random numbers between 1 and 38 in a list
#this code generate 10 lists
randomlist = []    # this list will contains the 10 random lists
def randomFun(numofProj,n):
    for k in range (0,n):
        randomlist1 = random.sample(range(1, numofProj), 36)
        randomlist.append(randomlist1)
    return randomlist

numofProject = 38    # the number of student projects in excel file
numOfParents = 10    # the number of random lists we generated
parentsRandom = []
parentsRandom = randomFun(numofProject,numOfParents)   # generate 10 lists
for j in range(len(parentsRandom)):
    print(parentsRandom[j])
#-----------------------------------------------------------------------------------------------------------------------
print("====================================TOTAL FITNESS FOR EACH RANDOM LIST====================================")

#random list 1
#open student excel file and read it
filename = openpyxl.load_workbook('Students+selections.xlsx')
allSheetNames = filename.sheetnames

fit1 = []
fit2 = []
fit3 = []
array = []
for k in range(len(randomlist)):
    fit1.append(array)
    fit2.append(array)
    fit3.append(array)
    fit1[k] = 0          # set fit1 & fit2 & fit3 Zero to find the fittness value for each random list
    fit2[k] = 0          # these three fit attribute is to store fittness pointS
    fit3[k] = 0          #fit1 --> store C values,fit2 --> store D values,fit3 --> store E values,

#To read all the selection students columns and compare each value with values in random list
for sheet in allSheetNames:
    currentSheet = filename[sheet]

    for row in range(2, currentSheet.max_row + 1): # it's range start from 2 that the first selection is in C2 Cell

        for column in "C":  # C is the first selection student columns
            cell_name = "{}{}".format(column, row)  # the first cell is C2 the range of row start from 2.
           # print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

            # we assume the indices of the random list are the groups id
            # [row-2] represents index in the random list
            # In the begining it will compare randomlist[i] to the value in C2 , C3 and so on
            for i in range(len(randomlist)):
                if ((randomlist[i][row - 2] == currentSheet[cell_name].value)):
                   fit1[i] += 30   # add 30 to the fit1 if there is duplicate

        for column in "D":  #CDE are the selection student columns
            cell_name = "{}{}".format(column, row) #the first cell is D2 the range of row start from 2.
            #print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

            # we assume the indexes of the random list are the groups id
            #[row-2] represents index in the random list
            # In the begening it will compare randomlist[i] to the value in D2, D3 and so on
            for i in range(len(randomlist)):
                if ((randomlist[i][row - 2] == currentSheet[cell_name].value)):
                    fit2[i] += 20  # add 20 to the fit2 if there is duplicate


        for column in "E":  #CDE are the selection student columns
            cell_name = "{}{}".format(column, row) #the first cell is E2 the range of row start from 2.
            #print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

            # we assume the indexes of the random list are the groups id
            #[row-2] represents index in the random list
            # In the begening it will compare randomlist[i] to the value in E2, E3 and so on
            for i in range(len(randomlist)):
                if ((randomlist[i][row - 2] == currentSheet[cell_name].value)):
                    fit3[i] += 10 # add 10 to the fit3 if there is duplicate

TotalFittness = []
for i in range(len(randomlist)):
    TotalFittness.append(array)
    TotalFittness[i] = fit1[i] + fit2[i] + fit3[i]  #summation of fi1, fit2, fit3 = total fitness of random list
    print('The total fittness of the random list  =  ', +TotalFittness[i])

print("====================================FIND THE 2 BEST RANDOM LIST====================================")
# add all the fittness values into list
TotalFittnessA = TotalFittness.copy() #copy() is to prevent the change on the TotalFittnessB
TotalFittnessB = TotalFittness.copy() #copy() is to prevent the change on the TotalFittnessA

# find the best random list - which has the highest fittness value -
# The best fitness for the best random value is 30 * 36 = 1080 which is the ideal case
max1 = max(TotalFittnessA)  # find the max value of the TotalFittness list
idx1 = TotalFittnessA.index(max1) # find the index of that max value
print('1) The  fitness of the first best = ',max1)  # print the first max value
print('The index of the first best in random list :',idx1) # print its index

TotalFittnessA.pop(idx1)   # pop that value
max2 = max(TotalFittnessA)  # find the max value after pop the max2
idx2 = TotalFittnessB.index(max2)  # find the index of max2
print('2) The  fitness of the second best = ',max2)
#print(idx2)

if max2 == max1:  #if the 2 first best are equal, print the index of each
    for i in range(idx1+1,len(TotalFittnessB)):
        if (max2 == TotalFittnessB[i]):
            idx2 = i

print('The index of the second best in random list :',idx2)
bestParent1 = []
bestParent2 = []
bestParent1 = randomlist[idx1]  # the first best random list
bestParent2 = randomlist[idx2]  # the second best random list

print("====================================THE FIRST & THE SECOND CHILD====================================")

# cross over function using single point cross over function
#cross over parent1 & parent2 to get child1 & child2
def Crossover(parent1,parent2):
    child11 = [[], []]   # To store the 2 children
    array1 = []
    for j in range(len(child11)):
        for i in range(0,len(parent1)):
            child11[j].append(array1)
    for j in range(len(child11)):
        if j == 0:
            parent3 = parent1
            parent4 = parent2
        else:
            if j == 1:
                parent3 = parent2
                parent4 = parent1
        index = 0
        index2 = int(len(child11[j])/2)  #integer is to prevent floating
        for k in range(0,int(len(child11[j])/2)):
                child11[j][k] = parent3[index]
                index += 1
        for m in range(int(len(child11[j])/2),len(child11[j])): #integer is to prevent floating
            if index2 != len(child11[j]):
                child11[j][m] = parent4[index2]
                index2 += 1
    return child11

child5 = []
child5 = Crossover(bestParent1,bestParent2)
print("The first child is :", child5[0])
print("The second child is :", child5[1])
print ("               -------------------------------------------------------------------------------------------------------                     ")

# mutation function to replace the duplicate the first & the second child
def Mutation(list, n):
    s = []
    n = len(list)  #n is the length of the child list
    for i in range(n):

    # check whether the element is repeated or not
        if list[i] not in s:
           s.append(list[i])
        else:
        # replace the duplicate element by another is not exist in the range 1-38
             for j in range(1, 38):
                 if j not in s:
                    list[i] = j
                    s.append(j)
                    break
    return s

# print the children after mutation
child5[0] = Mutation(child5[0], len(child5[0]))
print('The First child after MUTATION : ', child5[0])
child5[1] = Mutation(child5[1],len(child5[1]))
print('The second child after MUTATION : ',child5[1])

print("====================================THE FITNESS OF THE CHILDREN AFTER MUTATION====================================")
def Fittness(allSheetNames,child1Final):
    fit1 = 0  #fit1 --> store C values,fit2 --> store D values,fit3 --> store E values,
    fit2 = 0
    fit3 = 0

    # To read all the selection students columns and compare each value with values in random list
    for sheet in allSheetNames:
        currentSheet = filename[sheet]

        for row in range(2, currentSheet.max_row + 1):

            for column in "C":  # CDE are the selection student columns
                cell_name = "{}{}".format(column, row)  # the first cell is C2 the range of row start from 2.
                # print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

                # we assume the indexes of the random list are the groups id
                # [row-2] represents index in the random list
                # In the begining it will compare child1Final to the value in C2,C3 and so on
                if ((child1Final[row - 2] == currentSheet[cell_name].value)):
                  fit1 += 30 # add 30 to the fit1 if there is duplicate


            for column in "D":  # CDE are the selection student columns
                cell_name = "{}{}".format(column, row)  # the first cell is D2 the range of row start from 2.
                # print(cell_name)
                # print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

                # we assume the indexes of the random list are the groups id
                # [row-2] represents index in the random list
                # In the begening it will compare child1Final to the value in D2,D3 and so on
                if ((child1Final[row - 2] == currentSheet[cell_name].value)):
                    fit2 += 20  # add 20 to the fit2 if there is duplicate

            for column in "E":  # CDE are the selection student columns
                cell_name = "{}{}".format(column, row)  # the first cell is C2 the range of row start from 2.
                # print("{}-->{}".format(cell_name, currentSheet[cell_name].value))

                # we assume the indexes of the random list are the groups id
                # [row-2] represents index in the random list
                # In the begening it will compare child1Final to the value E2, E3 and so on
                if ((child1Final[row - 2] == currentSheet[cell_name].value)):
                    fit3 += 10  # add 10 to the fit3 if there is duplicate

    TotalFittness = 0
    TotalFittness = fit1 + fit2 + fit3   #summation of fi1, fit2, fit3 = total fitness of child list
    #print('The total fittness of the Child list  =  ', +TotalFittness)
    return TotalFittness

TotalFittnessCh = [[],[]]
for k in range(0,2):

    TotalFittnessCh[k]= Fittness(allSheetNames,child5[k])
    print('The total fittness of the Child list  =  ', +TotalFittnessCh[k])

print("====================================Cross the two child to get better result====================================")
child6 = []
child6 = Crossover(child5[0], child5[1])
print("The third child is :", child6[0])
print("The fourth child is :", child6[1])
print('------------------Childern list after mutation----------------------')
child6[0] = Mutation(child6[0], len(child6[0]))
print('The Third child after MUTATION : ', child6[0])
child6[1] = Mutation(child6[1],len(child6[1]))
print('The forth child after MUTATION : ',child6[1])

TotalFittnessCh1 = [[],[]]
for k in range(0,2):

    TotalFittnessCh1[k]= Fittness(allSheetNames,child6[k])
    print('The total fittness of the Child list  =  ', +TotalFittnessCh1[k])

print("====================================Cross the 5th and 6th child to get better result====================================")
child7 = []
child7 = Crossover(child6[0], child6[1])
print("The fifth child is :", child7[0])
print("The sixth child is :", child7[1])
print('------------------Childern list after mutation----------------------')
child6[0] = Mutation(child7[0], len(child7[0]))
print('The fifth child after MUTATION : ', child7[0])
child6[1] = Mutation(child7[1], len(child7[1]))
print('The sixth child after MUTATION : ', child7[1])

TotalFittnessCh2 = [[], []]
for k in range(0, 2):
    TotalFittnessCh2[k] = Fittness(allSheetNames, child6[k])
    print('The total fittness of the Child list  =  ', +TotalFittnessCh2[k])

TheSolutions = [TotalFittnessCh,TotalFittnessCh1,TotalFittnessCh2]
maxx = max(max(TheSolutions)) #find the highest fitness of the best fitness
print("The best fitness (solution) is the fittness : ", maxx) #print the solution fitness

